#!/usr/bin/env python3
"""
Relatório semanal — Cabine Zero 15
Roda 1x por semana via GitHub Actions.

O que faz:
1. Confere o certificado SSL do domínio (dias até vencer).
2. Roda uma checagem rápida de tempo de resposta em todas as páginas,
   pra achar a mais lenta do momento.
3. Lê data/incidents_week.csv (o que aconteceu durante a semana).
4. Monta um Excel com 3 abas: Resumo, Incidentes, Tempo de resposta.
5. Manda esse Excel por e-mail.
6. Zera data/incidents_week.csv — a semana seguinte começa do zero.
"""
import csv
import os
import smtplib
import ssl
import socket
import datetime
import urllib.request
import time
from email.message import EmailMessage

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

URLS_FILE = "urls.txt"
INCIDENTS_FILE = "data/incidents_week.csv"
THREAD_ID_FILE = "data/thread_id.txt"
DOMAIN = "www.cabinezero15.com.br"
TIMEOUT = 15

EMAIL_FROM = os.environ.get("EMAIL_FROM", "")
EMAIL_APP_PASSWORD = os.environ.get("EMAIL_APP_PASSWORD", "")
EMAIL_TO = os.environ.get("EMAIL_TO", "")


def check_ssl_days_left(domain):
    try:
        ctx = ssl.create_default_context()
        with socket.create_connection((domain, 443), timeout=TIMEOUT) as sock:
            with ctx.wrap_socket(sock, server_hostname=domain) as ssock:
                cert = ssock.getpeercert()
        expire_str = cert["notAfter"]
        expire_date = datetime.datetime.strptime(expire_str, "%b %d %H:%M:%S %Y %Z")
        days_left = (expire_date - datetime.datetime.utcnow()).days
        return days_left, expire_date.strftime("%d/%m/%Y")
    except Exception as e:
        return None, f"erro: {e}"


def measure_response_times(urls):
    times = {}
    for url in urls:
        start = time.time()
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "CabineZero15-LinkMonitor/2.0"})
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                resp.read()
                times[url] = round((time.time() - start) * 1000)
        except Exception:
            times[url] = None
    return times


def load_incidents():
    if not os.path.exists(INCIDENTS_FILE):
        return []
    with open(INCIDENTS_FILE, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_excel(ssl_days, ssl_expire_str, response_times, incidents):
    wb = Workbook()

    header_fill = PatternFill(start_color="1E1E1E", end_color="1E1E1E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # --- Aba 1: Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    total_pages = len(response_times)
    valid_times = [t for t in response_times.values() if t is not None]
    avg_time = round(sum(valid_times) / len(valid_times)) if valid_times else 0
    slowest_url = max(response_times, key=lambda u: (response_times[u] or -1))
    open_incidents = [i for i in incidents if i.get("resolved_at", "") == ""]

    rows = [
        ("Data do relatório", datetime.date.today().strftime("%d/%m/%Y")),
        ("Total de páginas monitoradas", total_pages),
        ("Incidentes na semana", len(incidents)),
        ("Incidentes ainda em aberto", len(open_incidents)),
        ("Tempo médio de resposta (ms)", avg_time),
        ("Página mais lenta", slowest_url),
        ("Tempo dessa página (ms)", response_times[slowest_url]),
        ("Certificado SSL válido até", ssl_expire_str),
        ("Dias restantes de SSL", ssl_days if ssl_days is not None else "erro ao verificar"),
    ]
    ws.append(["Métrica", "Valor"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50

    # --- Aba 2: Incidentes ---
    ws2 = wb.create_sheet("Incidentes")
    ws2.append(["URL", "Motivo", "Aberto em", "Resolvido em"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for i in incidents:
        ws2.append([i["url"], i["reason"], i["opened_at"], i.get("resolved_at", "")])
    for col, width in zip("ABCD", [55, 45, 20, 20]):
        ws2.column_dimensions[col].width = width

    # --- Aba 3: Tempo de resposta ---
    ws3 = wb.create_sheet("Tempo de resposta")
    ws3.append(["URL", "Tempo (ms)"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for url, t in sorted(response_times.items(), key=lambda kv: (kv[1] or -1), reverse=True):
        ws3.append([url, t if t is not None else "falhou"])
    ws3.column_dimensions["A"].width = 55
    ws3.column_dimensions["B"].width = 15

    filename = f"relatorio-cz15-{datetime.date.today().isoformat()}.xlsx"
    wb.save(filename)
    return filename


def send_email_with_attachment(filepath):
    if not EMAIL_FROM or not EMAIL_APP_PASSWORD or not EMAIL_TO:
        print("[aviso] Variáveis de e-mail não configuradas — pulando envio.")
        return

    msg = EmailMessage()
    msg["Subject"] = f"Relatório semanal Cabine Zero 15 — {datetime.date.today().strftime('%d/%m/%Y')}"
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.set_content(
        "Segue em anexo o relatório semanal de monitoramento do site.\n\n"
        "Este e-mail foi gerado automaticamente."
    )

    with open(filepath, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(filepath),
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_FROM, EMAIL_APP_PASSWORD)
        smtp.send_message(msg)
    print(f"E-mail enviado para {EMAIL_TO}.")


def reset_incidents_file():
    os.makedirs("data", exist_ok=True)
    with open(INCIDENTS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["url", "reason", "opened_at", "resolved_at"])
        writer.writeheader()


def reset_thread_id():
    """Reinicia a conversa de e-mail dos alertas — cada semana começa um thread novo."""
    if os.path.exists(THREAD_ID_FILE):
        os.remove(THREAD_ID_FILE)


def main():
    with open(URLS_FILE, encoding="utf-8") as f:
        urls = [line.strip() for line in f if line.strip()]

    print("Verificando certificado SSL...")
    ssl_days, ssl_expire_str = check_ssl_days_left(DOMAIN)

    print("Medindo tempo de resposta de todas as páginas...")
    response_times = measure_response_times(urls)

    print("Lendo incidentes da semana...")
    incidents = load_incidents()

    print("Montando Excel...")
    filename = build_excel(ssl_days, ssl_expire_str, response_times, incidents)

    print("Enviando e-mail...")
    send_email_with_attachment(filename)

    print("Zerando o arquivo de incidentes da semana...")
    reset_incidents_file()

    print("Reiniciando o thread de e-mail dos alertas para a nova semana...")
    reset_thread_id()

    print("Concluído.")


if __name__ == "__main__":
    main()
